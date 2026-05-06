# -*- coding: utf-8 -*-
import openpyxl, requests
from collections import defaultdict

URL = 'https://nlhhktqhupqnxnjxwqzd.supabase.co'
KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im5saGhrdHFodXBxbnhuanh3cXpkIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3Nzk5NzI1MiwiZXhwIjoyMDkzNTczMjUyfQ.WmMiO-3RATmCydfs74WhIPvtRZkMWjCi17ZMltIW7n0'

headers = {
    'Authorization': f'Bearer {KEY}',
    'apikey': KEY,
    'Content-Type': 'application/json',
    'Prefer': 'resolution=merge-duplicates'
}

def main():
    print('Reading order Excel...')
    wb = openpyxl.load_workbook('../jingxuan_orders_2026050602293354.xlsx')
    ws = wb.active
    
    accounts = defaultdict(lambda: {'operator': '', 'type': '小店', 'daily': defaultdict(int), 'totalIncome': 0, 'totalAmount': 0, 'totalNetIncome': 0})
    skipped_refund = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        account_raw = row[11]
        pay_time = row[3]
        operator = row[15] or ''
        account_type = row[16] or ''
        order_status = row[6] or ''
        
        # Skip refunds
        if order_status and '退款' in str(order_status):
            skipped_refund += 1
            continue
        
        try:
            income = float(row[10] or 0)
        except:
            income = 0
        try:
            amount = float(row[7] or 0)
        except:
            amount = 0
        
        if not account_raw or not pay_time:
            continue
        
        account = str(account_raw).split('(')[0].strip()
        date = str(pay_time)[:10]
        
        accounts[account]['operator'] = operator or accounts[account]['operator']
        accounts[account]['type'] = account_type or accounts[account]['type']
        accounts[account]['daily'][date] += 1
        accounts[account]['totalIncome'] += income
        accounts[account]['totalAmount'] += amount
        accounts[account]['totalNetIncome'] += income * 0.9
    
    # Read meta
    try:
        wb_meta = openpyxl.load_workbook('../抖音账号基础信息.xlsx')
        for row in wb_meta.active.iter_rows(min_row=2, values_only=True):
            name, typ, status, buyer = row
            if name and name in accounts:
                accounts[name]['buyer'] = buyer or ''
                accounts[name]['status'] = status or ''
                accounts[name]['type'] = typ or '混剪'
    except Exception as e:
        print('Meta read error:', e)
    
    print(f'Found {len(accounts)} accounts, skipped {skipped_refund} refunds')
    
    # Clear old data first
    print('Clearing old data...')
    requests.delete(f'{URL}/rest/v1/daily_stats', headers={**headers, 'Prefer': 'count=exact'}, params={'id': 'gt.0'})
    
    # Insert accounts
    account_payload = []
    for acc, data in accounts.items():
        account_payload.append({
            'name': acc,
            'operator': data['operator'],
            'account_type': data['type'],
            'buyer': data.get('buyer', ''),
            'status': data.get('status', '')
        })
    
    print('Inserting accounts...')
    for chunk in [account_payload[i:i+500] for i in range(0, len(account_payload), 500)]:
        res = requests.post(f'{URL}/rest/v1/accounts', headers={**headers, 'Prefer': 'resolution=merge-duplicates,return=minimal'}, json=chunk)
        print('Accounts batch:', res.status_code)
    
    # Insert stats
    stats = []
    for acc, data in accounts.items():
        for date, orders in data['daily'].items():
            stats.append({
                'account_name': acc,
                'date': date,
                'orders': orders,
                'income': round(data['totalIncome'], 2),
                'amount': round(data['totalAmount'], 2),
                'net_income': round(data['totalNetIncome'], 2)
            })
    
    print(f'Inserting {len(stats)} daily records...')
    for chunk in [stats[i:i+1000] for i in range(0, len(stats), 1000)]:
        res = requests.post(f'{URL}/rest/v1/daily_stats', headers={**headers, 'Prefer': 'resolution=merge-duplicates,return=minimal'}, json=chunk)
        print('Stats batch:', res.status_code)
    
    # Insert costs
    print('Reading cost Excel...')
    try:
        wb_cost = openpyxl.load_workbook('../消耗示例.xlsx')
        ws_cost = wb_cost.active
        cost_headers = [cell.value for cell in ws_cost[1]]
        
        from datetime import datetime, timedelta
        date_cols = []
        for i, h in enumerate(cost_headers[1:], 1):
            if isinstance(h, int):
                date = datetime(1899, 12, 30) + timedelta(days=h)
                date_cols.append({'idx': i, 'date': date.strftime('%Y-%m-%d')})
        
        costs = []
        for row in ws_cost.iter_rows(min_row=2, values_only=True):
            account_name = str(row[0] or '').strip()
            if not account_name:
                continue
            for dc in date_cols:
                cost = float(row[dc['idx']] or 0)
                if cost > 0:
                    costs.append({
                        'account_name': account_name,
                        'date': dc['date'],
                        'cost': round(cost, 2)
                    })
        
        print(f'Inserting {len(costs)} cost records...')
        for chunk in [costs[i:i+1000] for i in range(0, len(costs), 1000)]:
            res = requests.post(f'{URL}/rest/v1/daily_costs', headers={**headers, 'Prefer': 'resolution=merge-duplicates,return=minimal'}, json=chunk)
            print('Costs batch:', res.status_code)
    except Exception as e:
        print('Cost import error:', e)
    
    print('Done!')

if __name__ == '__main__':
    main()
